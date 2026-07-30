#!/usr/bin/env python3
"""
Video Catalog.xlsx  →  prisma/seed-data/learn-catalog.json   (brief_learn_v1 WS1)

Mirrors the Service Catalog pattern exactly: the spreadsheet is the author's
working document and stays OUT of git (23 MB); the generated JSON is what the
repo carries and what `seed-learn.ts` reads.

    python3 scripts/generate-learn-catalog.py

WHAT MAKES THIS FILE MESSY, and what is done about each:

  MERGED CELLS. The hierarchy columns are merged down over their children, so
  openpyxl reports the value on the top-left cell and `None` everywhere else.
  Merged ranges are expanded FIRST, before any forward-fill — the two are not
  interchangeable. Forward-fill alone would also carry a value across a genuine
  gap, and expanding alone would miss the continuation rows that are blank
  without being merged. Both, in that order.

  \\xa0 AS "EMPTY". The Finance and EndUser SCE sheets use a non-breaking space
  where the others leave a cell blank. Treated as blank; otherwise every one of
  those rows inherits "\\xa0" as its section title.

  SHIFTING HEADERS. Columns are matched BY NAME with per-sheet aliases
  (`Course` / `Course Name`, `Section` / `Section(s)`, `Lesson (Name)` /
  `Video Name`, `Style` / `Type`, `Learning Path` / `Primary Learning Path`),
  never by index — Finance and EndUser SCE insert two extra columns.

  NO VIDEO URLS. Verified across every cell of the workbook: the six content
  sheets carry a `Vimeo` HEADER with no values, and the only links live in
  `Archive Scott`, an older YouTube set that name-joins to 0 of 523 lessons.
  So `vimeo_ref` is null for every lesson, per the brief's decision block, and
  Archive Scott is not read. 296 lessons will still carry
  production_status = URL_ADDED_TO_LESSON: that flag is production TRACKING, not
  a URL, and WS2 gates playback on the URL rather than on the status.
"""
import json
import os
import re
import sys
from collections import Counter

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip3 install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
XLSX = os.path.join(HERE, "data", "Video Catalog.xlsx")
OUT = os.path.join(REPO, "prisma", "seed-data", "learn-catalog.json")

SHEETS = ["Beginners", "ERP", "Users - HCM", "Finance", "EndUser SCE", "Implementers"]

# Column aliases, in preference order.
COLS = {
    "audience": ["Course Type"],
    "group": ["Group"],
    "path": ["Learning Path", "Primary Learning Path"],
    "course": ["Course", "Course Name"],
    "style": ["Style", "Type"],
    "section": ["Section", "Section(s)"],
    "lesson": ["Lesson (Name)", "Video Name"],
    "description": ["Description", "Content Description"],
    "run_time": ["Run Time"],
    "expert": ["Expert"],
    "status": ["Status"],
}

AUDIENCE = {
    "beginners": "BEGINNERS",
    "end-user": "END_USER",
    "end-user-course": "END_USER",
    "implementer": "IMPLEMENTER",
    "content creator": "CONTENT_CREATOR",
}

# Only the values that genuinely correspond to a CourseStyle member. "OC
# Overiew" (sic) and "Conceptual" have no equivalent in the WS0 enum and are
# left null rather than forced into the nearest-looking one — inventing a
# mapping here would silently mislabel 77 courses. Reported as warnings.
STYLE = {
    "fa overview": "FA_OVERVIEW",
    "how to use": "HOW_TO_USE",
    "use": "HOW_TO_USE",
    "how to deploy": "HOW_TO_DEPLOY",
    "daily journal": "DAILY_JOURNAL",
    "ask the expert": "ASK_THE_EXPERT",
}

# The catalog's 0→7 ladder, keyed on the leading number so a reworded label
# still lands correctly. "0. Previously Recorded/Video Needs Refresh" is the
# spreadsheet's "0b" and is matched on its text.
STATUS_BY_NUM = {
    "0": "IN_CONCEPT",
    "1": "DECK_READY",
    "2": "RAW_SHOT",
    "3": "PRODUCED",
    "4": "LOADED_TO_STREAMING",
    "5": "URL_ADDED_TO_LESSON",
    "6": "BLOG_CREATED",
    "7": "BLOG_RELEASED",
}

# Rows that are layout, not content.
SKIP_LESSON = re.compile(r"^\s*(total run time|total|n/?a)\s*:?\s*$", re.I)

warnings = Counter()


def clean(v):
    """Cell → trimmed string, with the sheets' several spellings of 'empty'."""
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ").strip()
    return "" if s in ("", "-", "--", "n/a", "N/A") else s


def expand_merges(ws):
    """Give every cell in a merged range the range's value."""
    for rng in list(ws.merged_cells.ranges):
        top_left = ws.cell(row=rng.min_row, column=rng.min_col).value
        if top_left is None:
            continue
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                yield (row, col, top_left)


def slugify(*parts):
    s = "-".join(p for p in parts if p)
    s = re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")
    return re.sub(r"-{2,}", "-", s)[:120] or "item"


def map_status(raw):
    if not raw:
        return "IN_CONCEPT"
    low = raw.lower()
    if "refresh" in low:
        return "NEEDS_REFRESH"
    m = re.match(r"\s*(\d)", raw)
    if m and m.group(1) in STATUS_BY_NUM:
        return STATUS_BY_NUM[m.group(1)]
    warnings[f"unmapped status: {raw}"] += 1
    return "IN_CONCEPT"


def map_style(raw):
    if not raw:
        return None
    key = raw.lower().strip()
    if key in STYLE:
        return STYLE[key]
    warnings[f"unmapped style (left null): {raw}"] += 1
    return None


def map_audience(raw):
    if not raw:
        return None
    key = raw.lower().strip()
    if key in AUDIENCE:
        return AUDIENCE[key]
    warnings[f"unmapped audience: {raw}"] += 1
    return None


def read_sheet(ws, sheet_name):
    """One sheet → flat lesson rows with the hierarchy resolved."""
    headers = {}
    for c in ws[2]:
        if isinstance(c.value, str) and c.value.strip():
            headers.setdefault(c.value.strip(), c.column)

    col = {}
    for field, aliases in COLS.items():
        for a in aliases:
            if a in headers:
                col[field] = headers[a]
                break

    missing = [f for f in ("path", "lesson") if f not in col]
    if missing:
        warnings[f"{sheet_name}: missing column(s) {missing} — sheet skipped"] += 1
        return []

    # Merged values, applied over a copy of the grid.
    overlay = {}
    for row, c, val in expand_merges(ws):
        overlay[(row, c)] = val

    def cell(r, field):
        c = col.get(field)
        if not c:
            return ""
        return clean(overlay.get((r, c), ws.cell(row=r, column=c).value))

    rows = []
    carry = {"audience": "", "group": "", "path": "", "course": "", "style": "", "section": ""}

    for r in range(3, ws.max_row + 1):
        lesson = cell(r, "lesson")

        # Forward-fill the hierarchy from the row above wherever this row is
        # blank. Done for EVERY row, including skipped ones, so a "Total run
        # time:" divider doesn't reset the context for the rows beneath it.
        for f in ("audience", "group", "path", "course", "style", "section"):
            v = cell(r, f)
            if v:
                carry[f] = v

        if not lesson or SKIP_LESSON.match(lesson):
            continue
        # A lesson can sit above the first Learning Path declaration — the author
        # left the LP cell empty for a whole block. In "Users - HCM" that is 21
        # real Core HR lessons, and the next LP declared below them (row 42,
        # "Benefits Admin") is a DIFFERENT topic, so looking ahead would file
        # them under the wrong path. The Group is the honest fallback: it is the
        # coarsest thing the sheet actually asserts about them. Counted, not
        # silent — dropping 21 lessons or misfiling them are both worse than
        # importing them one level up and saying so.
        path_title = carry["path"] or carry["group"] or sheet_name
        if not carry["path"]:
            warnings[
                f"{sheet_name}: no Learning Path declared — fell back to "
                f"'{path_title}'"
            ] += 1

        rows.append(
            {
                "sheet": sheet_name,
                "audience": carry["audience"],
                "group": carry["group"],
                "path": path_title,
                # The catalog sometimes collapses LP = Course; when the Course
                # cell is empty the path's own title stands in, which is what
                # WS0's "allow an LP with exactly one Course" is for.
                "course": carry["course"] or carry["path"],
                "style": carry["style"],
                "section": carry["section"] or "Lessons",
                "lesson": lesson,
                "description": cell(r, "description"),
                "run_time": cell(r, "run_time"),
                "expert": cell(r, "expert"),
                "status": cell(r, "status"),
            }
        )
    return rows


def build():
    if not os.path.exists(XLSX):
        sys.exit(f"missing workbook: {XLSX}\nPlace 'Video Catalog.xlsx' in scripts/data/.")

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    flat = []
    for name in SHEETS:
        if name not in wb.sheetnames:
            warnings[f"sheet not found: {name}"] += 1
            continue
        flat.extend(read_sheet(wb[name], name))

    # --- nest: path → course → section → lesson, order of first appearance ---
    paths = {}
    for row in flat:
        audience = map_audience(row["audience"]) or "END_USER"
        # The LP slug carries audience+group because path TITLES repeat across
        # audiences ("Basic Procurement" exists more than once) and the slug is
        # globally unique.
        p_key = (audience, row["group"], row["path"])
        p = paths.get(p_key)
        if not p:
            p = paths[p_key] = {
                "title": row["path"],
                "slug": slugify(audience, row["group"], row["path"]),
                "audience": audience,
                "group": row["group"] or None,
                "sheet": row["sheet"],
                "courses": {},
            }
        c = p["courses"].get(row["course"])
        if not c:
            c = p["courses"][row["course"]] = {
                "title": row["course"],
                "slug": slugify(row["course"]),
                "style": map_style(row["style"]),
                "sections": {},
            }
        s = c["sections"].get(row["section"])
        if not s:
            s = c["sections"][row["section"]] = {"title": row["section"], "lessons": []}

        # (section, title) is the lesson's natural key, so a repeated title
        # inside one section is the same lesson listed twice.
        if any(l["title"] == row["lesson"] for l in s["lessons"]):
            warnings[f"duplicate lesson within a section: {row['lesson'][:48]}"] += 1
            continue

        s["lessons"].append(
            {
                "title": row["lesson"],
                "description": row["description"] or None,
                "run_time": row["run_time"] or None,
                # Null everywhere, by decision — see the module docstring.
                "vimeo_ref": None,
                "production_status": map_status(row["status"]),
                "expert": row["expert"] or None,
            }
        )

    out_paths = []
    for i, p in enumerate(paths.values()):
        courses = []
        for j, c in enumerate(p["courses"].values()):
            sections = []
            for k, s in enumerate(c["sections"].values()):
                sections.append(
                    {
                        "title": s["title"],
                        "sort_order": k,
                        "lessons": [
                            {**l, "sort_order": n} for n, l in enumerate(s["lessons"])
                        ],
                    }
                )
            courses.append(
                {
                    "title": c["title"],
                    "slug": c["slug"],
                    "style": c["style"],
                    "sort_order": j,
                    "sections": sections,
                }
            )
        out_paths.append(
            {
                "title": p["title"],
                "slug": p["slug"],
                "audience": p["audience"],
                "group": p["group"],
                "sheet": p["sheet"],
                "sort_order": i,
                "courses": courses,
            }
        )

    n_courses = sum(len(p["courses"]) for p in out_paths)
    n_sections = sum(len(c["sections"]) for p in out_paths for c in p["courses"])
    n_lessons = sum(
        len(s["lessons"]) for p in out_paths for c in p["courses"] for s in c["sections"]
    )
    n_urls = sum(
        1
        for p in out_paths
        for c in p["courses"]
        for s in c["sections"]
        for l in s["lessons"]
        if l["vimeo_ref"]
    )
    n_status5 = sum(
        1
        for p in out_paths
        for c in p["courses"]
        for s in c["sections"]
        for l in s["lessons"]
        if l["production_status"] in ("URL_ADDED_TO_LESSON", "BLOG_CREATED", "BLOG_RELEASED")
    )

    doc = {
        "_source": "4. Project Documents/2. Design/7. Video Catalog/Video Catalog.xlsx",
        "_generated_by": "scripts/generate-learn-catalog.py (brief_learn_v1 WS1)",
        "_note": (
            "Learn curriculum: LearningPath -> Course -> Section -> Lesson. "
            "vimeo_ref is null on EVERY lesson: the six content sheets carry a Vimeo "
            "header with no values, and the only links live in 'Archive Scott', an "
            "older YouTube set that name-joins to 0 of 523 lessons. Re-export with the "
            "Vimeo column populated and re-run the generator + seeder to fill them in "
            "— the seeder matches on the hierarchy path and UPDATES."
        ),
        "_counts": {
            "learning_paths": len(out_paths),
            "courses": n_courses,
            "sections": n_sections,
            "lessons": n_lessons,
            "lessons_with_vimeo_ref": n_urls,
            "lessons_at_status_url_added_or_beyond": n_status5,
        },
        "_warnings": [f"{k} (x{v})" for k, v in warnings.most_common()],
        "paths": out_paths,
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(doc, fh, indent=2, ensure_ascii=False)
        fh.write("\n")

    print(f"wrote {os.path.relpath(OUT, REPO)}")
    for k, v in doc["_counts"].items():
        print(f"  {k:42} {v}")
    if doc["_warnings"]:
        print("  warnings:")
        for w in doc["_warnings"]:
            print(f"    - {w}")


if __name__ == "__main__":
    build()
